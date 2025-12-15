"""
参数优化模块

提供不同的优化方法来寻找策略的最优参数组合:
- 网格搜索 (Grid Search)
- 随机搜索 (Random Search)
- 贝叶斯优化 (Bayesian Optimization)
- 遗传算法 (Genetic Algorithm)
"""

import os
import time
import json
import itertools
import random
import numpy as np
import pandas as pd
import matplotlib
# 设置matplotlib使用Agg后端，确保可以在无图形界面环境下正常绘图
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from typing import Dict, List, Any, Optional, Union, Callable, Tuple
from datetime import datetime

try:
    # Excel导出相关库
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from joblib import Parallel, delayed
    JOBLIB_AVAILABLE = True
except ImportError:
    JOBLIB_AVAILABLE = False
    
try:
    from skopt import gp_minimize
    from skopt.space import Real, Integer, Categorical
    SKOPT_AVAILABLE = True
except ImportError:
    SKOPT_AVAILABLE = False

def _convert_numpy_types(obj):
    """递归转换嵌套数据结构中的NumPy数据类型为Python原生类型
    
    Args:
        obj: 任意Python对象
        
    Returns:
        转换后的对象
    """
    # 处理NumPy标量类型
    if isinstance(obj, (np.int8, np.int16, np.int32, np.int64)):
        return int(obj)
    elif isinstance(obj, (np.float16, np.float32, np.float64)):
        return float(obj)
    elif isinstance(obj, (np.bool_)):
        return bool(obj)
    # 处理NumPy数组
    elif isinstance(obj, np.ndarray):
        return _convert_numpy_types(obj.tolist())
    # 处理字典
    elif isinstance(obj, dict):
        return {key: _convert_numpy_types(value) for key, value in obj.items()}
    # 处理列表或元组
    elif isinstance(obj, (list, tuple)):
        return [_convert_numpy_types(item) for item in obj]
    # 其他类型直接返回
    return obj

class ParameterOptimizer:
    """参数优化器"""
    
    def __init__(self, backtester, strategy, initialize=None, logger=None, strategy_name=None):
        """初始化参数优化器
        
        Args:
            backtester: 回测器实例
            strategy: 策略函数
            initialize: 初始化函数
            logger: 日志函数
            strategy_name: 策略名称，用于保存结果的文件夹名，如果为None则尝试从strategy函数名获取
        """
        self.backtester = backtester
        self.strategy = strategy
        self.initialize = initialize
        self.results = {}
        self.best_params = None
        self.best_result = None
        self.optimization_metric = 'sharpe_ratio'  # 默认优化指标
        
        # 保存评估过的所有参数结果
        self.all_evaluated_params = []
        
        # 优化开始时间和状态追踪
        self.start_time = None
        self.total_combinations = 0
        self.completed_combinations = 0
        
        # 日志函数
        self.logger = logger or print
        
        # 获取策略名称
        if strategy_name is not None:
            # 提取基本策略名称（如果包含优化方法，去掉）
            base_name = strategy_name.split('_')[0] if '_' in strategy_name else strategy_name
            self.strategy_name = base_name
        elif hasattr(strategy, '__name__'):
            self.strategy_name = strategy.__name__
        else:
            self.strategy_name = "unknown_strategy"
        
        # 保存优化方法名称（由具体优化方法设置）
        self.optimization_method = None
        
        # 创建结果目录
        self.results_dir = "optimization"
        os.makedirs(self.results_dir, exist_ok=True)
        
        # 保存当前优化时间戳
        self.timestamp = None
        
    def set_optimization_metric(self, metric, higher_is_better=True):
        """设置优化指标
        
        Args:
            metric: 优化指标，如'sharpe_ratio', 'total_return', 'max_drawdown'等
            higher_is_better: 是否越高越好
            
        Returns:
            self: 支持链式调用
        """
        self.optimization_metric = metric
        self.higher_is_better = higher_is_better
        return self
        
    def grid_search(self, param_grid, parallel=False, n_jobs=-1, progress_log_interval=5, skip_final_report=False):
        """网格搜索优化
        
        Args:
            param_grid: 参数网格，如{'fast_ma': [5, 10, 15], 'slow_ma': [20, 30, 40]}
            parallel: 是否并行计算
            n_jobs: 并行作业数，-1表示使用所有CPU
            progress_log_interval: 进度日志间隔(秒)
            skip_final_report: 是否跳过最终完整报告生成
            
        Returns:
            最优参数组合和对应的回测结果
        """
        # 设置当前优化方法
        self.optimization_method = "Grid"
        self.start_time = time.time()
        
        # 设置环境变量禁用图表和报告生成
        old_no_visual = os.environ.get('NO_VISUALIZATION', '')
        old_no_console = os.environ.get('NO_CONSOLE_LOG', '')
        os.environ['NO_VISUALIZATION'] = 'True'
        os.environ['NO_CONSOLE_LOG'] = 'True'
        
        try:
            # 生成所有参数组合
            param_names = list(param_grid.keys())
            param_values = list(param_grid.values())
            combinations = list(itertools.product(*param_values))
            self.total_combinations = len(combinations)
            
            self.logger(f"开始网格搜索，总参数组合数: {self.total_combinations}")
            self.logger(f"优化指标: {self.optimization_metric}, 越{'高' if self.higher_is_better else '低'}越好")
            self.logger("已禁用详细日志和可视化输出，优化过程更加高效")
            
            last_log_time = time.time()
            self.completed_combinations = 0
            
            if parallel and JOBLIB_AVAILABLE and n_jobs != 1:
                # 使用joblib进行并行计算
                self.logger(f"使用并行计算，作业数: {n_jobs}")
                
                results = Parallel(n_jobs=n_jobs)(
                    delayed(self._evaluate_params_wrapper)(
                        {param_names[i]: comb[i] for i in range(len(param_names))}
                    ) for comb in combinations
                )
                
                # 处理并行结果
                for params, metric_value, performance in results:
                    # 保存参数和结果
                    param_key = str(params)
                    if metric_value is not None:  # 只处理非None结果
                        self.results[param_key] = {
                            'params': params,
                            'metric_value': metric_value,
                            'performance': performance
                        }
                        
                        # 更新最优结果
                        if self.best_result is None or self._is_better(metric_value, self.best_result):
                            self.best_params = params.copy()
                            self.best_result = metric_value
                            self.logger(f"找到更好的参数: {params}, {self.optimization_metric}: {metric_value}")
                        
                        # 添加到全局评估列表
                        eval_result = {
                            'params': params,
                            'metric_value': metric_value,
                            'performance': performance
                        }
                        self.all_evaluated_params.append(eval_result)
                
                # 更新进度
                self.completed_combinations = self.total_combinations
                self._log_progress(force=True)
                
            else:
                # 顺序计算
                for i, comb in enumerate(combinations):
                    params = {param_names[j]: comb[j] for j in range(len(param_names))}
                    metric_value, performance = self._evaluate_params(params)
                    
                    # 更新进度
                    self.completed_combinations += 1
                    
                    # 定期记录进度
                    current_time = time.time()
                    if current_time - last_log_time > progress_log_interval:
                        self._log_progress()
                        last_log_time = current_time
            
            # 记录最终结果
            self._log_progress(force=True)
            
            # 保存优化结果
            self._save_optimization_results()
            
            # 为最优参数运行一次完整回测
            if self.best_params:
                self.logger(f"为最优参数运行完整回测: {self.best_params}")
                
                # 决定是否需要生成完整报告
                run_full_report = not skip_final_report
                
                # 只有在需要时才恢复环境变量，运行完整回测
                if run_full_report:
                    # 恢复环境变量，运行完整回测
                    os.environ.pop('NO_VISUALIZATION', None)
                    os.environ.pop('NO_CONSOLE_LOG', None)
                    
                    # 设置回原来的值（如果有）
                    if old_no_visual:
                        os.environ['NO_VISUALIZATION'] = old_no_visual
                    if old_no_console:
                        os.environ['NO_CONSOLE_LOG'] = old_no_console
                    
                    full_results = self.backtester.run(
                        strategy=self.strategy,
                        initialize=self.initialize,
                        strategy_params=self.best_params,
                        silent_mode=False  # 生成完整报告
                    )
                    
                    # 保存最优参数和结果
                    self._save_best_results(full_results)
                    
                    return self.best_params, full_results
                else:
                    # 不运行完整报告，只返回最优参数和简单结果
                    # 从已有的评估结果中获取最优参数的性能指标
                    best_performance = {}
                    if self.best_params:
                        param_key = str(self.best_params)
                        if param_key in self.results:
                            best_performance = self.results[param_key].get('performance', {})
                    
                    # 确保包含关键性能指标
                    simple_results = {
                        'performance': {
                            'optimization_metric': self.optimization_metric,
                            'best_value': self.best_result,
                            # 添加其他关键指标，确保至少包含sharpe_ratio
                            'sharpe_ratio': best_performance.get('sharpe_ratio', 0),
                            'total_return': best_performance.get('total_return', 0),
                            'max_drawdown': best_performance.get('max_drawdown', 0),
                            'win_rate': best_performance.get('win_rate', 0)
                        }
                    }
                    
                    # 确保数据类型正确
                    simple_results = _convert_numpy_types(simple_results)
                    return _convert_numpy_types(self.best_params), simple_results
            else:
                self.logger("警告: 未找到有效的最优参数")
                return None, None
                
        finally:
            # 确保即使发生异常也恢复环境变量
            os.environ.pop('NO_VISUALIZATION', None)
            os.environ.pop('NO_CONSOLE_LOG', None)
            
            # 设置回原来的值（如果有）
            if old_no_visual:
                os.environ['NO_VISUALIZATION'] = old_no_visual
            if old_no_console:
                os.environ['NO_CONSOLE_LOG'] = old_no_console
            
            # 恢复回测器的优化模式
            self.backtester.set_optimization_mode(False)
    
    def random_search(self, param_space, n_iter=10, parallel=False, n_jobs=-1, progress_log_interval=5, skip_final_report=False):
        """随机搜索优化
        
        Args:
            param_space: 参数空间，如{'fast_ma': (5, 20), 'slow_ma': (20, 50)}
            n_iter: 随机迭代次数
            parallel: 是否并行计算
            n_jobs: 并行作业数，-1表示使用所有CPU
            progress_log_interval: 进度日志间隔(秒)
            skip_final_report: 是否跳过最终完整报告生成
            
        Returns:
            最优参数组合和对应的回测结果
        """
        # 设置当前优化方法
        self.optimization_method = "Random"
        self.start_time = time.time()
        self.total_combinations = n_iter
        self.completed_combinations = 0
        
        # 设置环境变量禁用图表和报告生成
        old_no_visual = os.environ.get('NO_VISUALIZATION', '')
        old_no_console = os.environ.get('NO_CONSOLE_LOG', '')
        os.environ['NO_VISUALIZATION'] = 'True'
        os.environ['NO_CONSOLE_LOG'] = 'True'
        
        try:
            self.logger(f"开始随机搜索，总迭代次数: {n_iter}")
            self.logger(f"优化指标: {self.optimization_metric}, 越{'高' if self.higher_is_better else '低'}越好")
            self.logger("已禁用详细日志和可视化输出，优化过程更加高效")
            
            # 生成随机参数组合
            param_combinations = []
            for _ in range(n_iter):
                params = {}
                for param_name, param_range in param_space.items():
                    if isinstance(param_range, tuple) and len(param_range) == 2:
                        # 连续值范围
                        start, end = param_range
                        if isinstance(start, int) and isinstance(end, int):
                            # 整数范围
                            params[param_name] = random.randint(start, end)
                        else:
                            # 浮点数范围
                            params[param_name] = random.uniform(start, end)
                    elif isinstance(param_range, list):
                        # 离散值列表
                        params[param_name] = random.choice(param_range)
                    else:
                        raise ValueError(f"不支持的参数空间格式: {param_name}: {param_range}")
                param_combinations.append(params)
            
            last_log_time = time.time()
            
            if parallel and JOBLIB_AVAILABLE and n_jobs != 1:
                # 使用joblib进行并行计算
                self.logger(f"使用并行计算，作业数: {n_jobs}")
                
                results = Parallel(n_jobs=n_jobs)(
                    delayed(self._evaluate_params_wrapper)(params) 
                    for params in param_combinations
                )
                
                # 处理并行结果
                for params, metric_value, performance in results:
                    # 保存参数和结果
                    param_key = str(params)
                    if metric_value is not None:  # 只处理非None结果
                        self.results[param_key] = {
                            'params': params,
                            'metric_value': metric_value,
                            'performance': performance
                        }
                        
                        # 更新最优结果
                        if self.best_result is None or self._is_better(metric_value, self.best_result):
                            self.best_params = params.copy()
                            self.best_result = metric_value
                            self.logger(f"找到更好的参数: {params}, {self.optimization_metric}: {metric_value}")
                        
                        # 添加到全局评估列表
                        eval_result = {
                            'params': params,
                            'metric_value': metric_value,
                            'performance': performance
                        }
                        self.all_evaluated_params.append(eval_result)
                
                # 更新进度
                self.completed_combinations = self.total_combinations
                self._log_progress(force=True)
                
            else:
                # 顺序计算
                for params in param_combinations:
                    metric_value, performance = self._evaluate_params(params)
                    
                    # 更新进度
                    self.completed_combinations += 1
                    
                    # 定期记录进度
                    current_time = time.time()
                    if current_time - last_log_time > progress_log_interval:
                        self._log_progress()
                        last_log_time = current_time
            
            # 记录最终结果
            self._log_progress(force=True)
            
            # 保存优化结果
            self._save_optimization_results()
            
            # 为最优参数运行一次完整回测
            if self.best_params:
                self.logger(f"为最优参数运行完整回测: {self.best_params}")
                
                # 决定是否需要生成完整报告
                run_full_report = not skip_final_report
                
                # 只有在需要时才恢复环境变量，运行完整回测
                if run_full_report:
                    # 恢复环境变量，运行完整回测
                    os.environ.pop('NO_VISUALIZATION', None)
                    os.environ.pop('NO_CONSOLE_LOG', None)
                    
                    # 设置回原来的值（如果有）
                    if old_no_visual:
                        os.environ['NO_VISUALIZATION'] = old_no_visual
                    if old_no_console:
                        os.environ['NO_CONSOLE_LOG'] = old_no_console
                    
                    full_results = self.backtester.run(
                        strategy=self.strategy,
                        initialize=self.initialize,
                        strategy_params=self.best_params,
                        silent_mode=False  # 生成完整报告
                    )
                    
                    # 保存最优参数和结果
                    self._save_best_results(full_results)
                    
                    return self.best_params, full_results
                else:
                    # 不运行完整报告，只返回最优参数和简单结果
                    # 从已有的评估结果中获取最优参数的性能指标
                    best_performance = {}
                    if self.best_params:
                        param_key = str(self.best_params)
                        if param_key in self.results:
                            best_performance = self.results[param_key].get('performance', {})
                    
                    # 确保包含关键性能指标
                    simple_results = {
                        'performance': {
                            'optimization_metric': self.optimization_metric,
                            'best_value': self.best_result,
                            # 添加其他关键指标，确保至少包含sharpe_ratio
                            'sharpe_ratio': best_performance.get('sharpe_ratio', 0),
                            'total_return': best_performance.get('total_return', 0),
                            'max_drawdown': best_performance.get('max_drawdown', 0),
                            'win_rate': best_performance.get('win_rate', 0)
                        }
                    }
                    
                    # 确保数据类型正确
                    simple_results = _convert_numpy_types(simple_results)
                    return _convert_numpy_types(self.best_params), simple_results
            else:
                self.logger("警告: 未找到有效的最优参数")
                return None, None
        
        finally:
            # 确保即使发生异常也恢复环境变量
            os.environ.pop('NO_VISUALIZATION', None)
            os.environ.pop('NO_CONSOLE_LOG', None)
            
            # 设置回原来的值（如果有）
            if old_no_visual:
                os.environ['NO_VISUALIZATION'] = old_no_visual
            if old_no_console:
                os.environ['NO_CONSOLE_LOG'] = old_no_console
            
            # 恢复回测器的优化模式
            self.backtester.set_optimization_mode(False)
    
    def bayesian_optimization(self, param_space, n_iter=10, progress_log_interval=5, skip_final_report=False):
        """贝叶斯优化
        
        Args:
            param_space: 参数空间，如{'fast_ma': (5, 20), 'slow_ma': (20, 50)}
            n_iter: 优化迭代次数
            progress_log_interval: 进度日志间隔(秒)
            skip_final_report: 是否跳过最终完整报告生成
            
        Returns:
            最优参数组合和对应的回测结果
        """
        # 设置当前优化方法
        self.optimization_method = "Bayesian"
        
        # 设置环境变量禁用图表和报告生成
        old_no_visual = os.environ.get('NO_VISUALIZATION', '')
        old_no_console = os.environ.get('NO_CONSOLE_LOG', '')
        os.environ['NO_VISUALIZATION'] = 'True'
        os.environ['NO_CONSOLE_LOG'] = 'True'
        
        try:
            if not SKOPT_AVAILABLE:
                self.logger("错误: 贝叶斯优化需要安装scikit-optimize库 (pip install scikit-optimize)")
                return None, None
                
            self.start_time = time.time()
            self.total_combinations = n_iter
            self.completed_combinations = 0
            
            self.logger(f"开始贝叶斯优化，总迭代次数: {n_iter}")
            self.logger(f"优化指标: {self.optimization_metric}, 越{'高' if self.higher_is_better else '低'}越好")
            self.logger("已禁用详细日志和可视化输出，优化过程更加高效")
            
            # 创建参数空间
            dimensions = []
            param_names = []
            
            for param_name, param_range in param_space.items():
                param_names.append(param_name)
                
                if isinstance(param_range, tuple) and len(param_range) == 2:
                    # 连续值范围
                    start, end = param_range
                    if isinstance(start, int) and isinstance(end, int):
                        # 整数范围
                        dimensions.append(Integer(start, end, name=param_name))
                    else:
                        # 浮点数范围
                        dimensions.append(Real(float(start), float(end), name=param_name))
                elif isinstance(param_range, list):
                    # 离散值列表
                    dimensions.append(Categorical(param_range, name=param_name))
                else:
                    raise ValueError(f"不支持的参数空间格式: {param_name}: {param_range}")
            
            # 定义目标函数
            last_log_time = time.time()
            
            def objective(x):
                # 创建参数字典
                params = {param_names[i]: x[i] for i in range(len(param_names))}
                
                # 评估参数
                metric_value, _ = self._evaluate_params(params)
                
                # 更新进度
                self.completed_combinations += 1
                
                # 记录进度
                nonlocal last_log_time
                current_time = time.time()
                if current_time - last_log_time > progress_log_interval:
                    self._log_progress()
                    last_log_time = current_time
                
                # 注意：贝叶斯优化是最小化问题，如果higher_is_better为True，需要取负值
                return -metric_value if self.higher_is_better else metric_value
            
            # 运行贝叶斯优化
            result = gp_minimize(
                objective,
                dimensions=dimensions,
                n_calls=n_iter,
                random_state=0
            )
            
            # 记录最终结果
            self._log_progress(force=True)
            
            # 保存优化结果
            self._save_optimization_results()
            
            # 为最优参数运行一次完整回测
            if self.best_params:
                self.logger(f"为最优参数运行完整回测: {self.best_params}")
                
                # 决定是否需要生成完整报告
                run_full_report = not skip_final_report
                
                # 只有在需要时才恢复环境变量，运行完整回测
                if run_full_report:
                    # 恢复环境变量，运行完整回测
                    os.environ.pop('NO_VISUALIZATION', None)
                    os.environ.pop('NO_CONSOLE_LOG', None)
                    
                    # 设置回原来的值（如果有）
                    if old_no_visual:
                        os.environ['NO_VISUALIZATION'] = old_no_visual
                    if old_no_console:
                        os.environ['NO_CONSOLE_LOG'] = old_no_console
                    
                    full_results = self.backtester.run(
                        strategy=self.strategy,
                        initialize=self.initialize,
                        strategy_params=self.best_params,
                        silent_mode=False  # 生成完整报告
                    )
                    
                    # 保存最优参数和结果
                    self._save_best_results(full_results)
                    
                    return self.best_params, full_results
                else:
                    # 不运行完整报告，只返回最优参数和简单结果
                    # 从已有的评估结果中获取最优参数的性能指标
                    best_performance = {}
                    if self.best_params:
                        param_key = str(self.best_params)
                        if param_key in self.results:
                            best_performance = self.results[param_key].get('performance', {})
                    
                    # 确保包含关键性能指标
                    simple_results = {
                        'performance': {
                            'optimization_metric': self.optimization_metric,
                            'best_value': self.best_result,
                            # 添加其他关键指标，确保至少包含sharpe_ratio
                            'sharpe_ratio': best_performance.get('sharpe_ratio', 0),
                            'total_return': best_performance.get('total_return', 0),
                            'max_drawdown': best_performance.get('max_drawdown', 0),
                            'win_rate': best_performance.get('win_rate', 0)
                        }
                    }
                    
                    # 确保数据类型正确
                    simple_results = _convert_numpy_types(simple_results)
                    return _convert_numpy_types(self.best_params), simple_results
            else:
                self.logger("警告: 未找到有效的最优参数")
                return None, None
                
        finally:
            # 确保即使发生异常也恢复环境变量
            os.environ.pop('NO_VISUALIZATION', None)
            os.environ.pop('NO_CONSOLE_LOG', None)
            
            # 设置回原来的值（如果有）
            if old_no_visual:
                os.environ['NO_VISUALIZATION'] = old_no_visual
            if old_no_console:
                os.environ['NO_CONSOLE_LOG'] = old_no_console
            
            # 恢复回测器的优化模式
            self.backtester.set_optimization_mode(False)
    
    def _evaluate_params_wrapper(self, params):
        """并行计算的包装函数"""
        try:
            # 运行回测
            results = self.backtester.run(
                strategy=self.strategy,
                initialize=self.initialize,
                strategy_params=params,
                silent_mode=True
            )
            
            # 提取优化指标
            performance = results.get('performance', {})
            metric_value = performance.get(self.optimization_metric)
            
            # 如果指标不存在或为None，设置为0（无效值）
            if metric_value is None:
                self.logger(f"警告: 参数 {params} 的{self.optimization_metric}为None")
                metric_value = 0  # 使用0代替-Infinity
                # 设置一个有意义的性能数据
                performance = {
                    'sharpe_ratio': 0,
                    'total_return': 0,
                    'max_drawdown': 0,
                    'win_rate': 0,
                    'invalid_params': True  # 标记无效参数
                }
            
            # 确保指标是数值类型
            try:
                metric_value = float(metric_value)
            except (TypeError, ValueError):
                self.logger(f"警告: 参数 {params} 的{self.optimization_metric}不是数值类型: {metric_value}")
                metric_value = 0  # 使用0代替-Infinity
                # 设置一个有意义的性能数据
                performance = {
                    'sharpe_ratio': 0,
                    'total_return': 0,
                    'max_drawdown': 0,
                    'win_rate': 0,
                    'invalid_params': True  # 标记无效参数
                }
                
            # 返回完整的结果信息
            return params, metric_value, performance
            
        except Exception as e:
            self.logger(f"评估参数 {params} 时出错: {str(e)}")
            metric_value = 0  # 使用0代替-Infinity
            # 设置一个有意义的性能数据
            performance = {
                'sharpe_ratio': 0,
                'total_return': 0,
                'max_drawdown': 0,
                'win_rate': 0,
                'error': str(e),  # 添加错误信息
                'invalid_params': True  # 标记无效参数
            }
            return params, metric_value, performance
    
    def _evaluate_params(self, params):
        """评估单个参数组合
        
        Args:
            params: 参数字典
            
        Returns:
            (metric_value, performance): 评估指标值和性能数据
        """
        try:
            # 设置回测器为优化模式
            self.backtester.set_optimization_mode(True)
            
            # 运行回测时设置silent_mode=True，不生成图表和报告
            # 环境变量已经在外层方法中设置，不再需要在这里设置和恢复
            results = self.backtester.run(
                strategy=self.strategy,
                initialize=self.initialize,
                strategy_params=params,
                silent_mode=True  # 关键参数，静默模式
            )
            
            # 提取优化指标
            performance = results.get('performance', {})
            metric_value = performance.get(self.optimization_metric)
            
            # 如果指标不存在或为None，返回0（无效值）
            if metric_value is None:
                self.logger(f"警告: 参数 {params} 的{self.optimization_metric}为None")
                metric_value = 0  # 使用0代替-Infinity
                # 设置一个有意义的性能数据
                performance = {
                    'sharpe_ratio': 0,
                    'total_return': 0,
                    'max_drawdown': 0,
                    'win_rate': 0,
                    'invalid_params': True  # 标记无效参数
                }
            
            # 确保指标是数值类型
            try:
                metric_value = float(metric_value)
            except (TypeError, ValueError):
                self.logger(f"警告: 参数 {params} 的{self.optimization_metric}不是数值类型: {metric_value}")
                metric_value = 0  # 使用0代替-Infinity
                # 设置一个有意义的性能数据
                performance = {
                    'sharpe_ratio': 0,
                    'total_return': 0,
                    'max_drawdown': 0,
                    'win_rate': 0,
                    'invalid_params': True  # 标记无效参数
                }
            
            # 即使是负值，也保存结果并更新最优参数
            # 保存参数和结果
            param_key = str(params)
            self.results[param_key] = {
                'params': params,
                'metric_value': metric_value,
                'performance': performance
            }
            
            # 保存到全局评估列表
            eval_result = {
                'params': params,
                'metric_value': metric_value,
                'performance': {
                    k: v for k, v in performance.items() 
                    if k in ['total_return', 'sharpe_ratio', 'max_drawdown', 'win_rate', 'invalid_params', 'error']
                }
            }
            self.all_evaluated_params.append(eval_result)
            
            # 更新最优结果
            self._update_best_result(params, metric_value)
            
            return metric_value, performance
            
        except Exception as e:
            self.logger(f"评估参数 {params} 时出错: {str(e)}")
            metric_value = 0  # 使用0代替-Infinity
            # 设置一个有意义的性能数据
            performance = {
                'sharpe_ratio': 0,
                'total_return': 0,
                'max_drawdown': 0,
                'win_rate': 0,
                'error': str(e),  # 添加错误信息
                'invalid_params': True  # 标记无效参数
            }
            
            # 保存参数和结果
            param_key = str(params)
            self.results[param_key] = {
                'params': params,
                'metric_value': metric_value,
                'performance': performance
            }
            
            # 保存到全局评估列表
            eval_result = {
                'params': params,
                'metric_value': metric_value,
                'performance': performance
            }
            self.all_evaluated_params.append(eval_result)
            
            return metric_value, performance
    
    def _update_best_result(self, params, metric_value):
        """更新最优结果
        
        Args:
            params: 参数字典
            metric_value: 评估指标值
        """
        if self.best_result is None or self._is_better(metric_value, self.best_result):
            self.best_params = params.copy()
            self.best_result = metric_value
            
            # 记录找到更好参数
            self.logger(f"找到更好的参数: {params}, {self.optimization_metric}: {metric_value}")
    
    def _is_better(self, new_value, current_best):
        """判断新值是否优于当前最佳值
        
        Args:
            new_value: 新指标值
            current_best: 当前最佳值
            
        Returns:
            bool: 新值是否更好
        """
        # 如果优化目标是越高越好
        if self.higher_is_better:
            return new_value > current_best
        else:
            return new_value < current_best
    
    def _log_progress(self, force=False):
        """记录优化进度
        
        Args:
            force: 是否强制记录，无视时间间隔
        """
        if self.total_combinations <= 0:
            return
            
        elapsed_time = time.time() - self.start_time
        percentage = self.completed_combinations / self.total_combinations * 100
        
        # 计算剩余时间
        if self.completed_combinations > 0:
            avg_time_per_combo = elapsed_time / self.completed_combinations
            remaining_combos = self.total_combinations - self.completed_combinations
            remaining_time = avg_time_per_combo * remaining_combos
            
            # 转换为友好的时间格式
            remaining_hours = int(remaining_time // 3600)
            remaining_minutes = int((remaining_time % 3600) // 60)
            remaining_seconds = int(remaining_time % 60)
            
            time_estimate = f"{remaining_hours}h {remaining_minutes}m {remaining_seconds}s"
        else:
            time_estimate = "计算中..."
        
        # 记录进度
        self.logger(f"优化进度: {self.completed_combinations}/{self.total_combinations} " + 
                  f"({percentage:.2f}%), 已用时间: {elapsed_time:.1f}s, 预计剩余: {time_estimate}")
    
    def _save_optimization_results(self):
        """保存优化结果到文件"""
        # 设置时间戳（如果未设置）
        if self.timestamp is None:
            self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
        # 创建策略特定子文件夹（只用基本名称）
        strategy_folder = f"{self.strategy_name}_{self.timestamp}"
        strategy_results_dir = os.path.join(self.results_dir, strategy_folder)
        os.makedirs(strategy_results_dir, exist_ok=True)
        
        # 设置文件路径（包含优化方法）
        method_suffix = f"_{self.optimization_method}" if self.optimization_method else ""
        filename = os.path.join(strategy_results_dir, f"optimization_results{method_suffix}.json")
        
        # 准备保存的数据
        data = {
            'timestamp': self.timestamp,
            'strategy_name': self.strategy_name,
            'optimization_method': self.optimization_method,
            'optimization_metric': self.optimization_metric,
            'higher_is_better': self.higher_is_better,
            'total_combinations': self.total_combinations,
            'completed_combinations': self.completed_combinations,
            'best_params': self.best_params,
            'best_result': self.best_result
        }
        
        # 如果参数组合过多，只保存最好的N个参数组合
        max_params_to_save = 1000  # 最多保存1000个参数组合
        
        all_params_to_save = self.all_evaluated_params
        if len(self.all_evaluated_params) > max_params_to_save:
            self.logger(f"参数组合数量过多({len(self.all_evaluated_params)}), 只保存最好的{max_params_to_save}个")
            # 按指标值排序，保存最好的N个
            all_params_to_save = sorted(
                self.all_evaluated_params,
                key=lambda x: x['metric_value'],
                reverse=self.higher_is_better
            )[:max_params_to_save]
            
        data['all_evaluated_params'] = all_params_to_save
        
        # 使用通用转换函数处理所有数据
        data = _convert_numpy_types(data)
        
        # 保存到JSON文件
        with open(filename, 'w') as f:
            json.dump(data, f, indent=4)
        
        # 修改为超链接格式的日志输出
        abs_path = os.path.abspath(filename)
        self.logger(f"优化结果已保存到:{abs_path}")
        
        # 导出Excel文件
        try:
            if not EXCEL_AVAILABLE:
                self.logger("警告: 未安装openpyxl库，无法导出Excel文件。请运行 'pip install openpyxl' 安装。")
                return
                
            # 创建Excel文件路径
            excel_filename = os.path.join(strategy_results_dir, f"optimization_results{method_suffix}.xlsx")
            
            # 转换参数组合和性能指标为DataFrame
            rows = []
            for p in all_params_to_save:
                row = {**p['params']}  # 参数值
                row['metric_value'] = p['metric_value']  # 优化指标值
                
                # 添加其他性能指标
                if 'performance' in p:
                    for k, v in p['performance'].items():
                        if k in ['sharpe_ratio', 'total_return', 'max_drawdown', 'win_rate']:
                            row[k] = v
                            
                # 添加是否为最优参数的标记
                is_best = True
                for param_name, param_value in self.best_params.items():
                    if p['params'].get(param_name) != param_value:
                        is_best = False
                        break
                row['is_best'] = is_best
                
                # 添加是否为无效参数
                row['is_invalid'] = p['performance'].get('invalid_params', False)
                
                rows.append(row)
                
            # 创建DataFrame
            df = pd.DataFrame(rows)
            
            # 排序：最优参数在最前面，然后按指标值排序
            df = df.sort_values(['is_best', 'metric_value'], ascending=[False, not self.higher_is_better])
            
            # 保存到Excel文件
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                # 参数优化结果表
                df.to_excel(writer, sheet_name='优化结果', index=False)
                
                # 添加摘要信息表
                summary_data = {
                    '属性': ['策略名称', '优化方法', '优化指标', '优化方向', '参数组合总数', '完成评估数量', 
                            '最优参数', '最优指标值', '生成时间'],
                    '值': [
                        self.strategy_name,
                        self.optimization_method,
                        self.optimization_metric,
                        '越高越好' if self.higher_is_better else '越低越好',
                        self.total_combinations,
                        self.completed_combinations,
                        str(self.best_params),
                        self.best_result,
                        self.timestamp
                    ]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='摘要', index=False)
                
                # 尝试应用样式（如果可能）
                try:
                    workbook = writer.book
                    results_sheet = writer.sheets['优化结果']
                    
                    # 设置列宽
                    for col in results_sheet.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = (max_length + 2) * 1.2
                        results_sheet.column_dimensions[column].width = adjusted_width
                    
                    # 高亮最优参数行
                    best_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    
                    # 设置表头样式
                    for cell in results_sheet[1]:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
                    
                    # 高亮最优行
                    for row_idx, row in enumerate(df.itertuples(), start=2):
                        if row.is_best:
                            for cell in results_sheet[row_idx]:
                                cell.fill = best_fill
                except Exception:
                    # 如果样式应用失败，忽略错误继续
                    pass
            
            # 修改为超链接格式的日志输出
            abs_excel_path = os.path.abspath(excel_filename)
            self.logger(f"优化结果已导出到Excel:{abs_excel_path}")
            
        except Exception as e:
            self.logger(f"导出Excel文件时出错: {str(e)}")
        
        # 生成参数分布图（包含优化方法）
        self._plot_parameter_distribution(strategy_results_dir)
    
    def _save_best_results(self, full_results):
        """保存最优参数的完整回测结果"""
        # 确保已有时间戳并已创建目录
        if self.timestamp is None:
            self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
        # 创建策略特定子文件夹（只用基本名称）
        strategy_folder = f"{self.strategy_name}_{self.timestamp}"
        strategy_results_dir = os.path.join(self.results_dir, strategy_folder)
        os.makedirs(strategy_results_dir, exist_ok=True)
        
        # 设置文件路径（包含优化方法）
        method_suffix = f"_{self.optimization_method}" if self.optimization_method else ""
        filename = os.path.join(strategy_results_dir, f"best_params{method_suffix}.json")
        
        # 准备保存的数据
        data = {
            'timestamp': self.timestamp,
            'strategy_name': self.strategy_name,
            'optimization_method': self.optimization_method,
            'optimization_metric': self.optimization_metric,
            'best_params': self.best_params,
            'best_result': self.best_result,
            'performance': full_results.get('performance', {})
        }
        
        # 使用通用转换函数处理所有数据
        data = _convert_numpy_types(data)
        
        # 保存到文件
        with open(filename, 'w') as f:
            json.dump(data, f, indent=4)
        
        # 修改为超链接格式的日志输出
        abs_path = os.path.abspath(filename)
        self.logger(f"最优参数结果已保存到:{abs_path}")
    
    def _plot_parameter_distribution(self, output_dir):
        """绘制参数分布图
        
        Args:
            output_dir: 输出目录
        """
        if not self.all_evaluated_params:
            return
            
        try:
            # 将评估结果转换为DataFrame
            df_raw = pd.DataFrame([
                {**p['params'], 'metric_value': p['metric_value'], 'invalid': p['performance'].get('invalid_params', False)} 
                for p in self.all_evaluated_params
            ])
            
            # 过滤掉无效的参数
            df = df_raw[~df_raw['invalid']].copy() if 'invalid' in df_raw.columns else df_raw[df_raw['metric_value'] != 0].copy()
            invalid_count = len(df_raw) - len(df)
            if invalid_count > 0:
                self.logger(f"图表中过滤掉 {invalid_count} 个无效参数组合")
                
            # 只有当有效参数存在时才继续绘图
            if len(df) == 0:
                self.logger("没有有效的参数组合可供绘图")
                return
            
            # 只保留参数列
            param_cols = [col for col in df.columns if col not in ['metric_value', 'invalid']]
            
            # 计算每个参数的最优分布
            fig, axes = plt.subplots(len(param_cols), 1, figsize=(10, 4 * len(param_cols)))
            if len(param_cols) == 1:
                axes = [axes]
                
            for i, param in enumerate(param_cols):
                ax = axes[i]
                
                # 检查参数类型
                if df[param].dtype in [np.int64, np.float64]:
                    # 数值型参数，绘制散点图
                    ax.scatter(df[param], df['metric_value'], alpha=0.6)
                    ax.set_xlabel(param)
                    ax.set_ylabel(self.optimization_metric)
                    
                    # 添加最优点
                    if self.best_params and param in self.best_params:
                        best_value = self.best_params[param]
                        ax.axvline(x=best_value, color='r', linestyle='--', alpha=0.7)
                        ax.text(
                            best_value, 
                            ax.get_ylim()[0] + (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.05,
                            f'最优: {best_value}',
                            color='r'
                        )
                        
                    # 添加有效数据点和无效数据点的统计
                    if invalid_count > 0:
                        ax.text(
                            0.02, 0.98,
                            f"有效参数: {len(df)}, 无效参数: {invalid_count}",
                            transform=ax.transAxes,
                            va='top',
                            ha='left',
                            bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.5)
                        )
                else:
                    # 类别型参数，绘制箱型图
                    ax.boxplot(
                        [df[df[param] == val]['metric_value'] for val in df[param].unique()],
                        labels=df[param].unique()
                    )
                    ax.set_xlabel(param)
                    ax.set_ylabel(self.optimization_metric)
                    
                    # 添加最优点
                    if self.best_params and param in self.best_params:
                        best_value = self.best_params[param]
                        best_idx = list(df[param].unique()).index(best_value) + 1
                        ax.scatter([best_idx], [self.best_result], color='r', s=100, marker='*')
                        ax.text(
                            best_idx, 
                            self.best_result,
                            f'最优: {best_value}',
                            color='r'
                        )
                        
                    # 添加有效数据点和无效数据点的统计
                    if invalid_count > 0:
                        ax.text(
                            0.02, 0.98,
                            f"有效参数: {len(df)}, 无效参数: {invalid_count}",
                            transform=ax.transAxes,
                            va='top',
                            ha='left',
                            bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.5)
                        )
            
            # 设置标题
            fig.suptitle(f"参数分布图 (优化指标: {self.optimization_metric}, {'越高越好' if self.higher_is_better else '越低越好'})")
            plt.tight_layout()
            
            # 保存图像（包含优化方法）
            method_suffix = f"_{self.optimization_method}" if self.optimization_method else ""
            plot_path = os.path.join(output_dir, f"param_distribution{method_suffix}.png")
            plt.savefig(plot_path)
            plt.close()
            
            # 修改为超链接格式的日志输出
            abs_path = os.path.abspath(plot_path)
            self.logger(f"参数分布图已保存到:{abs_path}")
            
        except Exception as e:
            self.logger(f"绘制参数分布图时出错: {str(e)}")
    
    def get_optimization_results(self):
        """获取优化结果
        
        Returns:
            包含所有参数组合和对应结果的字典
        """
        results = {
            'best_params': self.best_params,
            'best_result': self.best_result,
            'all_results': self.results,
            'all_evaluated_params': self.all_evaluated_params
        }
        # 确保返回的所有数据都经过类型转换处理
        return _convert_numpy_types(results)
        
    def plot_optimization_results(self, save_path=None):
        """可视化优化结果
        
        Args:
            save_path: 图表保存路径，如果为None则使用策略结果目录
            
        Returns:
            图表路径列表
        """
        if not self.all_evaluated_params:
            self.logger("没有优化结果可以绘制")
            return []
            
        plot_paths = []
        
        # 如果未指定保存路径，使用策略特定目录
        if save_path is None:
            if self.timestamp is None:
                self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            strategy_folder = f"{self.strategy_name}_{self.timestamp}"
            save_path = os.path.join(self.results_dir, strategy_folder)
            os.makedirs(save_path, exist_ok=True)
        
        try:
            # 转换为DataFrame
            df = pd.DataFrame([
                {**p['params'], 'metric_value': p['metric_value']} 
                for p in self.all_evaluated_params
            ])
            
            # 按优化指标排序
            df = df.sort_values('metric_value', ascending=not self.higher_is_better)
            
            # 只保留数值型参数
            param_cols = [col for col in df.columns if col != 'metric_value']
            
            # 1. 绘制参数重要性图
            if len(param_cols) > 1:
                plt.figure(figsize=(10, 6))
                
                # 计算每个参数的相关性
                correlations = df[param_cols + ['metric_value']].corr()['metric_value'].drop('metric_value')
                correlations = correlations.abs().sort_values(ascending=False)
                
                bars = plt.barh(
                    range(len(correlations)), 
                    correlations.values,
                    tick_label=correlations.index
                )
                
                # 添加数值标签
                for bar in bars:
                    width = bar.get_width()
                    plt.text(
                        width + 0.01, 
                        bar.get_y() + bar.get_height()/2, 
                        f'{width:.3f}',
                        va='center'
                    )
                
                plt.xlabel('相关性(绝对值)')
                plt.ylabel('参数')
                plt.title('参数重要性(与优化指标的相关性)')
                plt.tight_layout()
                
                # 保存图表
                importance_path = os.path.join(save_path, f"param_importance.png")
                plt.savefig(importance_path)
                plt.close()
                plot_paths.append(importance_path)
                
                # 修改为超链接格式的日志输出
                abs_path = os.path.abspath(importance_path)
                self.logger(f"参数重要性图已保存到:{abs_path}")
            
            # 2. 绘制最优参数的散点图矩阵
            if len(param_cols) > 1:
                from pandas.plotting import scatter_matrix
                
                # 选取前20%的结果
                top_results = df.head(max(1, len(df) // 5))
                
                plt.figure(figsize=(12, 10))
                scatter_matrix(
                    top_results[param_cols], 
                    alpha=0.5, 
                    figsize=(12, 10), 
                    diagonal='kde'
                )
                plt.suptitle('前20%最优参数的分布矩阵')
                plt.tight_layout()
                plt.subplots_adjust(top=0.95)
                
                # 保存图表
                matrix_path = os.path.join(save_path, f"param_matrix.png")
                plt.savefig(matrix_path)
                plt.close()
                plot_paths.append(matrix_path)
                
                # 修改为超链接格式的日志输出
                abs_path = os.path.abspath(matrix_path)
                self.logger(f"参数矩阵图已保存到:{abs_path}")
            
            # 3. 绘制优化过程图（如果有迭代顺序）
            plt.figure(figsize=(10, 6))
            
            # 按评估顺序排序
            df_sorted = pd.DataFrame([
                {**p['params'], 'metric_value': p['metric_value'], 'idx': i} 
                for i, p in enumerate(self.all_evaluated_params)
            ]).sort_values('idx')
            
            # 计算累积最优值
            if self.higher_is_better:
                df_sorted['best_so_far'] = df_sorted['metric_value'].cummax()
            else:
                df_sorted['best_so_far'] = df_sorted['metric_value'].cummin()
            
            plt.plot(df_sorted['idx'], df_sorted['metric_value'], 'o-', alpha=0.3, label='当前值')
            plt.plot(df_sorted['idx'], df_sorted['best_so_far'], 'r-', label='最优值')
            plt.xlabel('迭代次数')
            plt.ylabel(self.optimization_metric)
            plt.title('优化过程')
            plt.legend()
            plt.grid(True, alpha=0.3)
            plt.tight_layout()
            
            # 保存图表
            process_path = os.path.join(save_path, f"optimization_process.png")
            plt.savefig(process_path)
            plt.close()
            plot_paths.append(process_path)
            
            # 修改为超链接格式的日志输出
            abs_path = os.path.abspath(process_path)
            self.logger(f"优化过程图已保存到:{abs_path}")
            
            self.logger(f"优化结果可视化已完成")
            
            return plot_paths
            
        except Exception as e:
            self.logger(f"绘制优化结果时出错: {str(e)}")
            return [] 